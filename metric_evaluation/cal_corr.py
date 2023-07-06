import json
import os
import pandas as pd
from scipy.stats import pearsonr, spearmanr, kendalltau, normaltest, shapiro
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import *
import numpy as np
import matplotlib.pyplot as plt


def cal_corr(metric1, metric2, method='kendalltau'):
    assert len(metric1) == len(metric2)
    assert len(metric1) == 14  # 14 models
    corr, p_value = None, None
    if method == 'pearsonr':
        corr, p_value = pearsonr(metric1, metric2)
    elif method == 'spearmanr':
        corr, p_value = spearmanr(metric1, metric2)
    elif method == 'kendalltau':
        corr, p_value = kendalltau(metric1, metric2)
    else:
        print('your specified correlation method: {} is not supported'.format(corr))
    return corr, p_value


def get_paired_rank(metric_score, metric):
    idx_list = []  # store the index of the model with bigger rating
    num_models = len(metric_score[metric])  # 14
    for i in range(num_models):  # 0, 1, ..., num_models-1
        for j in range(num_models - i - 1):  # 0, 1, ..., num_models-2-i
            model1_score = metric_score[metric][i]
            model2_score = metric_score[metric][j + i + 1]  # i+1, i+2, ..., num_models-1
            idx = i if model1_score > model2_score else j + i + 1
            idx_list.append(idx)
    return idx_list


def calculate_paired_acc(metric_score, metric, label_metric):
    metricA_rank = get_paired_rank(metric_score, metric)
    metricB_rank = get_paired_rank(metric_score, label_metric)
    acc = (np.array(metricA_rank) == np.array(metricB_rank)).sum() / len(metricB_rank)
    return acc


def get_sys_corr():
    f = open('scores.jsonl', 'r')
    overall_metric = {}
    for line in tqdm(f):
        inst = json.loads(line)
        model_names = list(inst.keys())
        model_names.remove('case')
        model_names.sort()
        for one_model in model_names:
            one_model_metric = inst[one_model]
            annot1 = one_model_metric['annot1']  # {'asp_rel': xxx, 'self_coh': xxx, 'sent_con': xxx, 'readblty': xxx}
            annot2 = one_model_metric['annot2']  # {'asp_rel': xxx, 'self_coh': xxx, 'sent_con': xxx, 'readblty': xxx}

            avr_annot = {key: (val + annot2[key]) / 2 for key, val in annot1.items()}

            one_model_metric.pop('annot1')
            one_model_metric.pop('annot2')
            # remove human annotations
            one_model_metric.update(avr_annot)

            if one_model not in overall_metric.keys():
                overall_metric[one_model] = one_model_metric
            else:
                overall_metric[one_model] = {key: val + overall_metric[one_model][key] for key, val
                                             in one_model_metric.items()}

    assert len(overall_metric) == 14

    # {'bart': {'metric1': xxx, 'metric2': xxx, ..., 'metricN': xxx}, 'model2': {...}, ...}
    sys_metric_scores = {metric: [] for metric in overall_metric['bart'].keys()}

    for mdl_name in overall_metric.keys():
        one_model_dict = overall_metric[mdl_name]
        for metric in one_model_dict.keys():
            score = one_model_dict[metric] / 100  # N=100
            sys_metric_scores[metric].append(score)

    if not os.path.exists('results'):
        os.mkdir('results')
    metric_p_value_dict = {}
    wb = Workbook()
    save_path = 'results/sys_corr.xlsx'
    worksheet = wb.active

    wb_p_value = Workbook()
    p_value_save_path = 'results/sys_p_value.xlsx'
    worksheet_p_value = wb_p_value.active

    # wb_acc = Workbook()
    # acc_save_path = 'results/paired_acc.xlsx'
    # worksheet_acc = wb_acc.active

    for metric in sys_metric_scores.keys():
        stat, p = shapiro(sys_metric_scores[metric])
        if p < 0.05:
            # print('{} is not normally distributed'.format(metric))
            print('')
        else:
            latex_line = '\\textbf{' + metric + '}' + '& {:.3f} & {:.3f}\\\\'.format(stat, p)
            print(latex_line)
        corr_asp, p_value_asp = cal_corr(metric1=sys_metric_scores[metric], metric2=sys_metric_scores['asp_rel'])
        corr_self, p_value_self = cal_corr(metric1=sys_metric_scores[metric], metric2=sys_metric_scores['self_coh'])
        corr_sent, p_value_sent = cal_corr(metric1=sys_metric_scores[metric], metric2=sys_metric_scores['sent_con'])
        corr_read, p_value_read = cal_corr(metric1=sys_metric_scores[metric], metric2=sys_metric_scores['readblty'])

        worksheet.append([metric, corr_asp, corr_self, corr_sent, corr_read])
        metric_p_value_dict[metric] = [p_value_asp, p_value_self, p_value_sent, p_value_read]

        worksheet_p_value.append([metric, p_value_asp, p_value_self, p_value_sent, p_value_read])

        # worksheet_acc.append([metric,
        #                       calculate_paired_acc(metric_score=sys_metric_scores, metric=metric,
        #                                            label_metric='asp_rel'),
        #                       calculate_paired_acc(metric_score=sys_metric_scores, metric=metric,
        #                                            label_metric='self_coh'),
        #                       calculate_paired_acc(metric_score=sys_metric_scores, metric=metric,
        #                                            label_metric='sent_con'),
        #                       calculate_paired_acc(metric_score=sys_metric_scores, metric=metric,
        #                                            label_metric='readblty')])

    for row in worksheet.iter_rows():
        metric_name = row[0].value
        # print(metric_name)
        for d, cell in enumerate(row[1:]):
            # print(d)
            # print(cell)
            # print(metric_p_value_dict[metric_name])
            if metric_p_value_dict[metric_name][d] <= 0.01:
                cell.fill = PatternFill('solid', fgColor=Color('ED7D31'))
            elif metric_p_value_dict[metric_name][d] <= 0.05:
                cell.fill = PatternFill('solid', fgColor=Color('F4B183'))
            else:
                continue
    wb.save(filename=save_path)
    wb_p_value.save(filename=p_value_save_path)
    # wb_acc.save(filename=acc_save_path)


def get_summ_corr():
    f = open('scores.jsonl', 'r')
    overall_metric = {}
    for inst_idx, line in enumerate(tqdm(f)):
        inst = json.loads(line)
        model_names = list(inst.keys())
        model_names.remove('case')
        model_names.sort()

        for one_model in model_names:
            one_model_metric = inst[one_model]

            annot1 = one_model_metric['annot1']  # {'asp_rel': xxx, 'self_coh': xxx, 'sent_con': xxx, 'readblty': xxx}
            annot2 = one_model_metric['annot2']  # {'asp_rel': xxx, 'self_coh': xxx, 'sent_con': xxx, 'readblty': xxx}

            avr_annot = {key: (val + annot2[key]) / 2 for key, val in annot1.items()}
            # avr_annot = {key: val / 2 for key, val in avr_annot.items()}

            one_model_metric.pop('annot1')
            one_model_metric.pop('annot2')
            # remove human annotations
            one_model_metric.update(avr_annot)

            for metric in one_model_metric.keys():
                if metric not in overall_metric.keys():
                    overall_metric[metric] = [[] for _ in range(100)]
                overall_metric[metric][inst_idx].append(one_model_metric[metric])
    assert len(overall_metric['R1_F1'][0]) == 14
    # {'R1_F1': [[14 scores for case 1], [14 scores for case 2], ...], 'metric2': [...], ...}

    if not os.path.exists('results'):
        os.mkdir('results')
    wb = Workbook()
    save_path = 'results/summ_corr.xlsx'
    worksheet = wb.active
    for metric in overall_metric.keys():
        summ_asp, summ_self, summ_sent, summ_read = [], [], [], []
        for i in range(100):
            corr_asp, _ = cal_corr(metric1=overall_metric[metric][i], metric2=overall_metric['asp_rel'][i])
            corr_self, _ = cal_corr(metric1=overall_metric[metric][i], metric2=overall_metric['self_coh'][i])
            corr_sent, _ = cal_corr(metric1=overall_metric[metric][i], metric2=overall_metric['sent_con'][i])
            corr_read, _ = cal_corr(metric1=overall_metric[metric][i], metric2=overall_metric['readblty'][i])
            summ_asp.append(corr_asp)
            summ_self.append(corr_self)
            summ_sent.append(corr_sent)
            summ_read.append(corr_read)
        worksheet.append(
            [metric, sum(summ_asp) / 100, sum(summ_self) / 100, sum(summ_sent) / 100, sum(summ_read) / 100])
    wb.save(filename=save_path)


get_sys_corr()
get_summ_corr()
