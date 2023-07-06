import json
from tqdm import tqdm
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np


def get_ROUGE(model_name):  # for Table 2
    f = open('autometric_scores/rouge/rouge_14model_outputs.jsonl', 'r')
    avr_r1, avr_r2, avr_rL = 0, 0, 0
    for line in tqdm(f):
        inst = json.loads(line)
        rouge_score = inst['model_output'][model_name]['ppl_score']
        r1 = rouge_score['R1_F1']
        r2 = rouge_score['R2_F1']
        rL = rouge_score['RL_F1']

        avr_r1 += r1
        avr_r2 += r2
        avr_rL += rL

    avr_r1 /= 100
    avr_r2 /= 100
    avr_rL /= 100
    f.close()
    print(avr_r1, avr_r2, avr_rL)

    f = open('autometric_scores/rouge/rouge_14model_outputs.jsonl', 'r')

    for line in tqdm(f):
        inst = json.loads(line)
        rouge_score = inst['model_output'][model_name]['ppl_score']
        r1 = rouge_score['R1_F1']
        r2 = rouge_score['R2_F1']
        rL = rouge_score['RL_F1']
        if r1 >= avr_r1 and r2 >= avr_r2 and rL >= avr_rL:
            print(inst['case'])
            print('reference: {}'.format(inst['summ']))
            print('model output: {}'.format(inst['model_output'][model_name]['model_summ']))
            print('R1: {}, R2: {}, RL: {}'.format(r1, r2, rL))


def get_human_rating():  # for Table 5
    f = open('scores.jsonl', 'r')
    model_human_score_dict = {}
    dimensions = ["asp_rel", "self_coh", "sent_con", "readblty"]
    for line in tqdm(f):
        inst = json.loads(line)
        for model_name in inst.keys():
            if model_name == 'case':
                continue
            if model_name not in model_human_score_dict.keys():
                model_human_score_dict[model_name] = {"asp_rel": [], "self_coh": [], "sent_con": [], "readblty": [],
                                                      'r1': [], 'r2': [], 'rL': [], 'bart_arti_hypo': []}
            # print(inst[model_name]['metric']['annot1'])
            model_annot1 = inst[model_name]['annot1']
            model_annot2 = inst[model_name]['annot2']
            model_human_score_dict[model_name]['r1'].append(inst[model_name]['R1_F1'])
            model_human_score_dict[model_name]['r2'].append(inst[model_name]['R2_F1'])
            model_human_score_dict[model_name]['rL'].append(inst[model_name]['RL_F1'])
            model_human_score_dict[model_name]['bart_arti_hypo'].append(inst[model_name]['bart_score_arti2hypo'])
            for d in dimensions:
                model_human_score_dict[model_name][d].append((model_annot1[d] + model_annot2[d]) / 2)
    for model_name in model_human_score_dict.keys():
        for d in model_human_score_dict[model_name].keys():
            assert len(model_human_score_dict[model_name][d]) == 100
            model_human_score_dict[model_name][d] = sum(model_human_score_dict[model_name][d]) / 100

    for model_name in model_human_score_dict.keys():
        human_score_dict = model_human_score_dict[model_name]
        latex_line = '\\textbf{' + model_name + '}' + '& {:.3f} & {:.3f} & {:.3f} & {:.3f} & {:.4f} & {:.4f} & {:.4f} & {:.3f}' \
                                                      '\\\\'.format(
            human_score_dict['asp_rel'],
            human_score_dict['self_coh'],
            human_score_dict['sent_con'],
            human_score_dict['readblty'],
            human_score_dict['r1'], human_score_dict['r2'], human_score_dict['rL'], human_score_dict['bart_arti_hypo'])
        print(latex_line)


def get_metric_eval_result():  # for Table 4
    wb_sys_corr = load_workbook('results/sys_corr.xlsx')
    wb_sys_p_val = load_workbook('results/sys_p_value.xlsx')
    wb_sum_corr = load_workbook('results/summ_corr.xlsx')
    for row_sys, row_p, row_summ in tqdm(zip(wb_sys_corr['Sheet'].iter_rows(),
                                             wb_sys_p_val['Sheet'].iter_rows(),
                                             wb_sum_corr['Sheet'].iter_rows())):
        metric_name = row_sys[0].value
        assert metric_name == row_p[0].value
        assert metric_name == row_summ[0].value

        latex_line = '\\textbf{' + metric_name + '}'
        for cell_sys, cell_p, cell_sum in zip(row_sys[1:], row_p[1:], row_summ[1:]):
            sys_val = cell_sys.value
            sys_p = cell_p.value
            sum_val = cell_sum.value
            if sys_p > 0.05:
                dim_latex_line = '{:.2f} & {:.5f}'.format(sys_val, sum_val)
            elif sys_p > 0.01:
                dim_latex_line = '{:.2f}$^*$ & {:.5f}'.format(sys_val, sum_val)
            else:
                dim_latex_line = '{:.2f}'.format(sys_val) + '$^{**}$ &' + ' {:.5f}'.format(sum_val)
            # xx & xx
            latex_line = latex_line + ' & ' + dim_latex_line
        print(latex_line + ' \\\\')


def plot_histogram():  # for Figure 1
    f = open('scores.jsonl', 'r')
    dimensions = ["asp_rel", "self_coh", "sent_con", "readblty"]
    avr_scores = []
    human_score_dict = {dim: [] for dim in dimensions}
    for line in tqdm(f):
        inst = json.loads(line)
        for model_name in inst.keys():
            if model_name == 'case':
                continue
            model_annot1 = inst[model_name]['annot1']
            model_annot2 = inst[model_name]['annot2']
            for dim in dimensions:
                score1, score2 = model_annot1[dim], model_annot2[dim]
                human_score_dict[dim].append(score1)
                human_score_dict[dim].append(score2)
                avr_scores.append((score1 + score2) / 2)
    max_distri = 0
    for dim in dimensions:
        one_dim_cnt = human_score_dict[dim]
        distri_cnt = {score + 1: 0 for score in range(5)}
        for score in one_dim_cnt:  # a list
            distri_cnt[score] += 1
        assert sum([distri_cnt[key] for key in distri_cnt.keys()]) == 100 * 14 * 2  # 2800
        for key in distri_cnt.keys():
            distri_cnt[key] /= 2800
        max_distri = max(max_distri, max([distri_cnt[key] for key in distri_cnt.keys()]))

        human_score_dict[dim] = [distri_cnt[score + 1] for score in range(5)]

    plt.figure(figsize=(6, 4))
    plt.margins(0, 0)
    width = 0.225
    x = np.array([1, 2, 3, 4, 5])

    avr_list, upper_list, lower_list = [], [], []
    for i in range(5):
        score = i + 1
        prob = []
        for dim in dimensions:
            prob.append(human_score_dict[dim][score - 1])
        avr_prob = sum(prob) / 4
        std = np.std(prob)
        print(prob, std)

        avr_list.append(avr_prob)
        upper_list.append(avr_prob + std)
        lower_list.append(avr_prob - std)
    plt.plot(x, avr_list, 'black', label='dimension average')
    plt.fill_between(x, upper_list, lower_list, facecolor='pink', edgecolor='red', alpha=0.5,
                     label='standard derivation')

    for i in range(4):
        dim = dimensions[i]
        data = human_score_dict[dim]
        print(dim, data)
        if i == 0:
            title = 'aspect relevance'
        elif i == 1:
            title = 'self-coherence'
        elif i == 2:
            title = 'sentiment consistency'
        else:
            title = 'readability'

        plt.bar(x + (-1.5 + i) * width, data, width, label=title, edgecolor='black')
    plt.scatter(x=x, y=upper_list, color='black', s=5, marker='v')
    plt.scatter(x=x, y=lower_list, color='black', s=5, marker='^')

    plt.legend(loc='best')

    plt.savefig('annot_distri.png', dpi=600, bbox_inches='tight')
    plt.show()


